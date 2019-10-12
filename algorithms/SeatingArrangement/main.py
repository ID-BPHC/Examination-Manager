import datetime
from algorithms.SeatingArrangement.Course import CourseList
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def get_dates_from_key(key):

    splitted = key.split("|")

    start = datetime.datetime.strptime(
        splitted[0].strip() + " " + splitted[1].strip(), "%d/%m/%y %H:%M")
    end = datetime.datetime.strptime(
        splitted[0].strip() + " " + splitted[2].strip(), "%d/%m/%y %H:%M")

    return (start, end)


def is_no_exam_code(code):

    no_exam_codes = ["F266", "F366", "F367", "F376", "F377"]

    for ne_code in no_exam_codes:
        if ne_code in code:
            return True

    if code[-1] == "T":
        return True

    return False


def get_course_list(file_name):

    f = open(file_name)

    course_list = CourseList()

    for line in f.readlines()[1:]:

        splitted = line.split(",")
        room = splitted[0]
        code = splitted[1]
        name = splitted[2]
        student_count = int(splitted[4])
        start, end = get_dates_from_key(splitted[6])

        course_list.add_if_not_exists(code, name, start, end)
        course = course_list.find_by_code(code)
        course.rooms.append((room, student_count))

    f.close()

    return course_list


def add_students(course_list, file_name):

    f = open(file_name)

    invalid_courses = set()

    for line in f.readlines():

        splitted = line.strip().split(",")
        student_id = splitted[0]
        course_code = splitted[1]
        course = course_list.find_by_code(course_code)

        if course is not None:
            if student_id not in course.students:
                course.students.append(student_id)

        else:
            if not is_no_exam_code(course_code):
                invalid_courses.add(course_code)

    f.close()

    return invalid_courses


def export_xlsx(course_list, file_name):

    wb = Workbook()
    ws = wb.active
    current_row = 1

    for course in course_list.courses:

        if len(course.students) == 0:
            print(
                f"No students enroled in {course.code}. Please re-check Registered Students list.")
            continue

        ws[f"A{current_row}"] = f"{course.code}"
        ws[f"B{current_row}"] = f"{course.name}"

        date_string = f"{course.exam_start.strftime('%d %B %y - %I:%M %p')} to {course.exam_end.strftime('%I:%M %p')}"
        ws[f"C{current_row}"] = date_string

        exam_slices = course.get_exam_slices()
        ws[f"D{current_row}"] = exam_slices[0][0]
        ws[f"E{current_row}"] = exam_slices[0][2] + " to " + exam_slices[0][3]
        ws[f"F{current_row}"] = exam_slices[0][1]

        current_row += 1

        for exam_slice in exam_slices[1:]:
            ws[f"D{current_row}"] = exam_slice[0]
            ws[f"E{current_row}"] = exam_slice[2] + " to " + exam_slice[3]
            ws[f"F{current_row}"] = exam_slice[1]
            current_row += 1

    wb.save(file_name)


def start_seating_arrangement_process(room_allotment_csv, students_csv):
    course_list = get_course_list(room_allotment_csv)

    invalid_courses = add_students(course_list, students_csv)

    print("The following courses do not have a room allotment (please re-confirm):\n")

    for course in invalid_courses:
        print(f"\t=> {course}")

    course_list.sort_entries()

    export_xlsx(
        course_list, "./SeatingArrangement.xlsx")
    
    print("Seating Arrangement exported to SeatingArrangement.xlsx")
