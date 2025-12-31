import json, sys
import os
import shutil
import openpyxl
import time
import csv

from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, PatternFill


class Course:
    def __init__(self, code, title):
        self.code = code
        self.title = title
        self.ic_email = ""
        self.rooms = []
        self.time = None
        self.students = []
        self.allotment_index = 0

    def get_next_student(self):
        if self.allotment_index >= len(self.students):
            return None

        else:
            student = self.students[self.allotment_index]
            self.allotment_index += 1
            return student


class CourseList:
    def __init__(self):
        self.courses = []

    def add_course(self, code, name):
        self.courses.append(Course(code, name))

    def add_if_not_exists(self, code, name):
        if self.find_by_code(code) is None:
            self.add_course(code, name)

    def find_by_code(self, code):
        for course in self.courses:
            if course.code == code:
                return course

            splitted = course.code.split("/")

            for splitted_course in splitted:
                if code.strip() == splitted_course.strip():
                    return course

        return None

    def sort_entries(self):
        for course in self.courses:
            course.students.sort()

    def __repr__(self):
        return f"{len(self.courses)} courses"


def get_matched_rooms(room_map, number):
    keys = []

    for key in room_map.keys():
        if number in key:
            keys.append(key)

    return keys


def get_populated_maps(
    room_map_csv, room_allotment_csv, registered_students_csv, ic_csv
):
    room_map = {}
    final_solution = {}
    course_list = CourseList()

    print("Reading Data")

    f = open(room_map_csv)

    for line in f.readlines():
        line = line.strip()
        splitted = line.split(",")
        room_number = splitted[0]
        no_of_cols = int(splitted[1])
        col_map = [int(x) for x in splitted[2 : no_of_cols + 2]]

        room_map[room_number] = col_map

    f.close()

    f = open(room_allotment_csv)

    for line in f.readlines()[1:]:
        line = line.strip()
        splitted = line.split(",")
        room = splitted[0]
        code = splitted[1]
        title = splitted[2]
        capacity = int(splitted[3])
        student_count = int(splitted[4])
        time = splitted[6]
        flag = splitted[7]

        course_list.add_if_not_exists(code, title)
        course = course_list.find_by_code(code)
        if course is not None:
            course.rooms.append((room, flag, student_count, capacity))
            course.time = time

        if time not in final_solution:
            final_solution[time] = {}

        if room not in final_solution[time]:
            matched_room_keys = get_matched_rooms(room_map, room)

            for key in matched_room_keys:
                seating_map = []

                for i in range(0, max(room_map[key])):
                    seating_map.append([])

                    for j in range(0, len(room_map[key])):
                        seating_map[i].append("")

                final_solution[time][key] = seating_map

    f.close()

    f = open(registered_students_csv)

    for line in f.readlines():
        line = line.strip()
        splitted = line.split(",")

        name = splitted[0]
        id_number = splitted[1]
        course_code = splitted[2]

        course = course_list.find_by_code(course_code)

        if course is not None:
            course.students.append(f"{id_number} - {name}")

    f.close()

    course_list.sort_entries()

    f = open(ic_csv)

    for line in f.readlines():
        line = line.strip()
        splitted = line.split(",")

        course_code = splitted[0]
        email = splitted[1]

        course = course_list.find_by_code(course_code)

        if course is not None:
            course.ic_email = email

    f.close()

    return room_map, final_solution, course_list


def generate_seating_charts(
    room_map_csv, room_allotment_csv, registered_students_csv, ic_csv
):
    print("Generating Seating Charts")
    room_map, final_solution, course_list = get_populated_maps(
        room_map_csv, room_allotment_csv, registered_students_csv, ic_csv
    )
    left_out_students = {}
    left_out_students_copy = {}
    not_alloted_students = 0
    for course in course_list.courses:
        for room, remark, student_count, capacity in course.rooms:
            keys = get_matched_rooms(room_map, room)
            seated = 0
            count = 0
            for key in keys:
                chart = final_solution[course.time][key]
                limits = room_map[key]

                half_cap = int(int(capacity / 2))
                start_point = 1 if student_count <= half_cap else 0
                step_value = 1 if remark == "FULL" else 2

                # Edit

                for i in range(0, len(limits)):
                    if chart[len(chart) - start_point - 1][i] != "":
                        start_point = start_point ^ 1
                    count += start_point
                    f = False  # flag to make sure first only the chessboard filling is done
                    if (
                        remark != "FULL" and student_count > half_cap
                    ):  # check if uneven fill
                        if count == capacity - student_count:
                            row = start_point
                        elif count > capacity - student_count:
                            row = 0
                        elif count < capacity - student_count:
                            for j in range(
                                start_point, limits[i], step_value
                            ):  # fill in chessboard pattern for maximum possible students
                                if count >= capacity - student_count:
                                    row = j
                                    # to continous fill from this point
                                    break
                                if j + step_value >= limits[i]:
                                    f = True
                                if chart[len(chart) - j - 1][i] == "":
                                    student = course.get_next_student()
                                    chart[len(chart) - j - 1][
                                        i
                                    ] = f"{course.code} - {student}"
                                    seated += 1
                                    if limits[i] - j - 2 >= 0:
                                        count += 1

                        if f:
                            start_point = start_point ^ 1
                            continue

                        while row <= limits[i]:

                            start_point = 0
                            if seated >= student_count:
                                break
                            if chart[len(chart) - row - 1][i] == "":
                                student = course.get_next_student()
                                chart[limits[i] - row - 1][
                                    i
                                ] = f"{course.code} - {student}"

                                seated += 1
                            else:
                                while chart[len(chart) - row - 1][i] != "":
                                    if row < limits[i]:
                                        row += 1
                                    else:
                                        break
                                if row < limits[i]:
                                    student = course.get_next_student()

                                    chart[len(chart) - row - 1][
                                        i
                                    ] = f"{course.code} - {student}"

                                    seated += 1
                            row += 1

                    else:
                        for j in range(start_point, limits[i], step_value):
                            if seated >= student_count:
                                break
                            if chart[len(chart) - j - 1][i] == "":
                                student = course.get_next_student()

                                chart[len(chart) - j - 1][
                                    i
                                ] = f"{course.code} - {student}"

                                seated += 1

                        if remark != "FULL":
                            start_point = start_point ^ 1
            # Edit

            for i in range(seated, student_count):
                key_dict = (room, course.time)
                if key_dict not in left_out_students:
                    left_out_students[key_dict] = {}
                    left_out_students_copy[key_dict] = {}
                if course.code not in left_out_students[key_dict]:
                    left_out_students[key_dict][course.code] = []
                    left_out_students_copy[key_dict][course.code] = []
                student = course.get_next_student()
                left_out_students[key_dict][course.code].append(student)
                left_out_students_copy[key_dict][course.code].append(student)
                not_alloted_students += 1

        if course.allotment_index < len(course.students):
            print(
                f"Seating Arrangement Discrepancy - {len(course.students) - course.allotment_index} students after {course.get_next_student()} for {course.code} - {course.title}"
            )

    left_out_students_count = 0
    for time_room, courses in left_out_students.items():
        room = time_room[0]
        for course_code in courses:
            count = 0
            course = course_list.find_by_code(course_code)
            keys = get_matched_rooms(room_map, room)
            for key in keys:
                chart = final_solution[course.time][key]
                limits = room_map[key]
                for i in range(0, len(limits)):
                    for j in range(start_point, limits[i], step_value):
                        if count == len(left_out_students[time_room][course_code]):
                            break

                        student = left_out_students[time_room][course_code][count]

                        if chart[limits[i] - j - 1][i] == "":
                            chart[limits[i] - j - 1][i] = f"{course.code} - {student}"
                            left_out_students_copy[time_room][course_code].remove(
                                student
                            )
                            count += 1
                            left_out_students_count += 1
    export_charts(room_map, course_list, final_solution)
    print(
        "Number of students alloted after alloting consecutive seats where required",
        left_out_students_count,
    )
    print(
        "Number of students which are still not alloted ",
        not_alloted_students - left_out_students_count,
    )

    print("Exporting Error file to error_file.csv ...")

    with open("error_file.csv", "w", newline="") as csv_file:
        writer = csv.writer(csv_file)
        for time_room, course_dict in left_out_students_copy.items():
            count = 0
            for course, students in course_dict.items():
                for student in students:
                    if count == 0:
                        writer.writerow([time_room[0], time_room[1], course, student])
                    else:
                        writer.writerow(["", "", "", student])
                    count += 1

    print("***** Done *****")


def export_charts(room_map, course_list, final_solution):
    print("***** Starting Chart Generation *****")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    heading_font = Font(size=13, bold=True)
    sub_heading_font = Font(size=11, bold=True)

    if os.path.exists("./Charts_and_Sheets") and os.path.isdir("./Charts_and_Sheets"):
        shutil.rmtree("./Charts_and_Sheets")
        time.sleep(0.5)

    os.mkdir("./Charts_and_Sheets")

    if os.path.exists("./Charts_and_Sheets") and os.path.isdir("./Charts_and_Sheets"):
        shutil.rmtree("./Charts_and_Sheets")
        time.sleep(0.5)

    os.mkdir("./Charts_and_Sheets")

    for course in course_list.courses:
        if course.time is None:
            continue

        if not os.path.exists(f"./Charts_and_Sheets/{course.ic_email}"):
            os.mkdir(f"./Charts_and_Sheets/{course.ic_email}")

        if not os.path.exists(f"./Charts_and_Sheets/{course.ic_email}"):
            os.mkdir(f"./Charts_and_Sheets/{course.ic_email}")

        wb_seating = openpyxl.Workbook()

        wb_attendance = openpyxl.Workbook()

        for room, flag, student_count, capacity in course.rooms:
            keys = get_matched_rooms(room_map, room)

            for key in keys:
                list_students = []
                print(f"Generating {course.code} - {key}")

                wb_seating.create_sheet(key)
                wb_attendance.create_sheet(key)
                ws_seating = wb_seating[key]
                ws_attendance = wb_attendance[key]
                ws_seating.sheet_properties.pageSetUpPr.fitToPage = True
                ws_seating.page_setup.fitToHeight = False
                openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(
                    ws_seating,
                    paper_size=ws_seating.PAPERSIZE_A4,
                    orientation="landscape",
                )

                ws_attendance.sheet_properties.pageSetUpPr.fitToPage = True
                ws_attendance.page_setup.fitToHeight = False
                openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(
                    ws_attendance,
                    paper_size=ws_attendance.PAPERSIZE_A4,
                    orientation="portrait",
                )

                total_cols = len(room_map[key])
                heading = f"{key} - {course.code} - {course.title}"

                end_char = chr(64 + total_cols)
                ws_seating.merge_cells(f"A1:{end_char}1")
                ws_seating.merge_cells(f"A2:{end_char}2")

                ws_attendance.merge_cells(f"A1:D1")
                ws_attendance.merge_cells(f"A2:D2")
                ws_attendance.merge_cells(f"A3:D3")

                ws_seating["A1"] = heading
                ws_seating["A1"].font = heading_font

                img = Image("./logo.png")

                img.width = 450
                img.height = 100
                img.left = 20
                ws_attendance.add_image(img, "B1")
                with open("./examHeading.txt") as f:
                    data = f.read()
                ws_attendance["A2"].font = heading_font
                ws_attendance["A2"] = data
                ws_attendance["A3"] = heading
                ws_attendance["A3"].font = heading_font
                ws_attendance["A4"] = "S.NO"
                ws_attendance["A4"].font = sub_heading_font
                ws_attendance["B4"] = "ID NUMBER"
                ws_attendance["B4"].font = sub_heading_font
                ws_attendance["C4"] = "NAME"
                ws_attendance["C4"].font = sub_heading_font
                ws_attendance["D4"] = "SIGNATURE"
                ws_attendance["D4"].font = sub_heading_font

                ws_seating["A2"] = "***** Blackboard Here *****"
                ws_seating["A2"].font = sub_heading_font
                ws_attendance.print_title_rows = "1:4"
                ws_attendance.print_options.horizontalCentered = True

                for row in final_solution[course.time][key]:
                    ws_seating.append(
                        [
                            (
                                "-".join([x.split("-")[0], x.split("-")[1]])
                                if x != ""
                                else x
                            )
                            for x in row
                        ]
                    )
                    for x in row:
                        if x.split("-")[0].strip() == course.code:
                            try:
                                list_students.append(
                                    x.split("-")[1] + "-" + x.split("-")[2]
                                )
                            except:
                                print(course.code, x)
                list_students.sort()
                for i, stud in enumerate(list_students):
                    ws_attendance.append(
                        [i + 1, stud.split("-")[0], stud.split("-")[1], ""]
                    )

                for col in range(65, 90):
                    ws_seating.column_dimensions[chr(col)].width = 15

                ws_attendance.row_dimensions[1].height = 20
                ws_attendance.column_dimensions["B"].width = 15
                for col in range(67, 90):
                    ws_attendance.column_dimensions[chr(col)].width = 35

                for row in ws_seating.iter_rows():
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(
                            wrap_text=True, horizontal="center", vertical="center"
                        )
                        cell.border = thin_border

                        if (
                            cell.value is not None
                            and cell.value.split("-")[0].strip() == course.code
                        ):
                            cell.fill = PatternFill(
                                start_color="E8E8E8",
                                end_color="E8E8E8",
                                fill_type="solid",
                            )

                for row in ws_attendance.iter_rows():
                    ws_attendance.row_dimensions[row[0].row].height = 25

                    for cell in row:
                        if cell.column == 3:
                            cell.alignment = openpyxl.styles.Alignment(
                                wrap_text=False, horizontal="left", vertical="center"
                            )
                        else:
                            cell.alignment = openpyxl.styles.Alignment(
                                wrap_text=False, horizontal="center", vertical="center"
                            )
                        if cell.row > 2:
                            cell.border = thin_border
                ws_attendance.row_dimensions[1].height = 80
                ws_attendance.row_dimensions[2].height = 40

        del wb_seating["Sheet"]
        del wb_attendance["Sheet"]

        try:
            path = os.path.join(
                "Charts_and_Sheets",
                course.ic_email,
                course.code.replace("/", "_") + " Seating Charts" + ".xlsx",
            )

            path1 = os.path.join(
                "Charts_and_Sheets",
                course.ic_email,
                course.code.replace("/", "_") + " Attendance Sheets" + ".xlsx",
            )

            wb_seating.save(path)
            wb_attendance.save(path1)
        except Exception as e:
            print(e)
            print("Could not create ", course.ic_email, " ", course.code)


if __name__ == "__main__":
    generate_seating_charts(
        r"C:\Users\Anirudh\Desktop\New folder\TTDdata\24-25 sem 1\compre\room_map.csv",
        r"C:\Users\Anirudh\Desktop\New folder\TTDdata\24-25 sem 1\compre\RoomAllotment 123.csv",
        r"C:\Users\Anirudh\Desktop\New folder\TTDdata\24-25 sem 1\compre\NAME ATTENDANCE STUDENT.csv",
        r"C:\Users\Anirudh\Desktop\New folder\TTDdata\24-25 sem 1\compre\IC DETAILS.csv",
    )
